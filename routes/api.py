"""API REST de inventario (productos, tipos, sync, export, IA)."""
import base64
import io
import json
import os
import uuid
from datetime import datetime, timedelta

from flask import Blueprint, Response, current_app, jsonify, request

from db import IntegrityError, get_db
from services.imagen import guardar_foto

api_bp = Blueprint('api', __name__)

gemini_client = None
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
if GEMINI_API_KEY:
    try:
        from google import genai
        gemini_client = genai.Client(api_key=GEMINI_API_KEY)
    except Exception as e:
        print(f"Gemini no disponible: {e}")

EXPORT_COLUMNAS = [
    ('nombre', 'Nombre'),
    ('tipo_nombre', 'Tipo'),
    ('cantidad', 'Cantidad'),
    ('precio_unitario', 'Precio unitario'),
    ('valor_total', 'Valor total'),
    ('codigo_barras', 'Código de barras'),
    ('descripcion', 'Descripción'),
    ('latitud', 'Latitud'),
    ('longitud', 'Longitud'),
    ('fecha_creacion', 'Fecha de creación'),
]


@api_bp.get('/api/config')
def get_config():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT moneda_simbolo FROM configuracion WHERE id = 1")
    row = cursor.fetchone()
    conn.close()
    if row:
        return jsonify({'moneda_simbolo': row['moneda_simbolo']})
    return jsonify({'moneda_simbolo': 'S/'})


@api_bp.post('/api/config')
def update_config():
    data = request.get_json(silent=True) or {}
    moneda_simbolo = (data.get('moneda_simbolo') or 'S/').strip()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE configuracion SET moneda_simbolo = ? WHERE id = 1", (moneda_simbolo,))
    conn.commit()
    conn.close()
    return jsonify({'moneda_simbolo': moneda_simbolo}), 200


@api_bp.get('/api/tipos-producto')
def get_tipos():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT t.*, COUNT(p.id) as num_productos
        FROM tipos_producto t
        LEFT JOIN productos p ON p.tipo_producto_id = t.id
        GROUP BY t.id
        ORDER BY t.nombre
    """)
    tipos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(tipos)


@api_bp.post('/api/tipos-producto')
def create_tipo():
    data = request.get_json(silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'El nombre del tipo es obligatorio'}), 400

    tipo_id = f"tipo-{uuid.uuid4().hex[:8]}"
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO tipos_producto (id, nombre, descripcion, icono, color)
            VALUES (?, ?, ?, ?, ?)
        """, (tipo_id, nombre, data.get('descripcion', ''),
              data.get('icono', '📦'), data.get('color', '#3b82f6')))
        conn.commit()
        cursor.execute("SELECT * FROM tipos_producto WHERE id = ?", (tipo_id,))
        tipo = dict(cursor.fetchone())
        conn.close()
        return jsonify(tipo), 201
    except IntegrityError:
        conn.close()
        return jsonify({'error': 'Ya existe ese tipo'}), 409


@api_bp.put('/api/tipos-producto/<tipo_id>')
def update_tipo(tipo_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM tipos_producto WHERE id = ?", (tipo_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'No encontrado'}), 404
    actual = dict(row)

    nombre = (data.get('nombre') or actual['nombre']).strip()
    if not nombre:
        conn.close()
        return jsonify({'error': 'El nombre del tipo es obligatorio'}), 400

    try:
        cursor.execute(
            "UPDATE tipos_producto SET nombre = ?, descripcion = ?, icono = ?, color = ? WHERE id = ?",
            (nombre,
             data.get('descripcion', actual['descripcion']),
             data.get('icono') or actual['icono'],
             data.get('color') or actual['color'],
             tipo_id)
        )
        conn.commit()
    except IntegrityError:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'Ya existe otro tipo con ese nombre'}), 409

    cursor.execute("SELECT * FROM tipos_producto WHERE id = ?", (tipo_id,))
    tipo = dict(cursor.fetchone())
    conn.close()
    return jsonify(tipo), 200


@api_bp.delete('/api/tipos-producto/<tipo_id>')
def delete_tipo(tipo_id):
    reasignar_a = request.args.get('reasignar_a')
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) as count FROM tipos_producto WHERE id = ?", (tipo_id,))
    if cursor.fetchone()['count'] == 0:
        conn.close()
        return jsonify({'error': 'No encontrado'}), 404

    cursor.execute("SELECT COUNT(*) as count FROM productos WHERE tipo_producto_id = ?", (tipo_id,))
    num_productos = cursor.fetchone()['count']

    if num_productos > 0:
        if not reasignar_a:
            conn.close()
            return jsonify({
                'error': 'El tipo tiene productos asociados',
                'productos': num_productos,
            }), 409
        if reasignar_a == tipo_id:
            conn.close()
            return jsonify({'error': 'El tipo destino no puede ser el mismo'}), 400
        cursor.execute("SELECT COUNT(*) as count FROM tipos_producto WHERE id = ?", (reasignar_a,))
        if cursor.fetchone()['count'] == 0:
            conn.close()
            return jsonify({'error': 'El tipo destino no existe'}), 400
        cursor.execute(
            "UPDATE productos SET tipo_producto_id = ? WHERE tipo_producto_id = ?",
            (reasignar_a, tipo_id)
        )

    cursor.execute("DELETE FROM tipos_producto WHERE id = ?", (tipo_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Eliminado', 'reasignados': num_productos}), 200


@api_bp.get('/api/productos')
def get_productos():
    tipo_id = request.args.get('tipo_id')
    busqueda = request.args.get('q')
    conn = get_db()
    cursor = conn.cursor()

    query = """
        SELECT p.*, t.nombre as tipo_nombre, t.icono as tipo_icono, t.color as tipo_color
        FROM productos p
        JOIN tipos_producto t ON p.tipo_producto_id = t.id
        WHERE 1=1
    """
    params = []
    if tipo_id:
        query += " AND p.tipo_producto_id = ?"
        params.append(tipo_id)
    if busqueda:
        query += " AND (p.nombre LIKE ? OR p.codigo_barras LIKE ?)"
        params.extend([f'%{busqueda}%', f'%{busqueda}%'])
    query += " ORDER BY p.fecha_creacion DESC"

    limit = request.args.get('limit', type=int)
    offset = request.args.get('offset', type=int) or 0
    if limit:
        query += " LIMIT ? OFFSET ?"
        params.extend([limit, offset])

    cursor.execute(query, params)
    productos = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(productos)


@api_bp.get('/api/productos/<producto_id>')
def get_producto(producto_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.*, t.nombre as tipo_nombre, t.icono as tipo_icono, t.color as tipo_color
        FROM productos p
        JOIN tipos_producto t ON p.tipo_producto_id = t.id
        WHERE p.id = ?
    """, (producto_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'No encontrado'}), 404
    return jsonify(dict(row))


@api_bp.post('/api/productos')
def create_producto():
    nombre = request.form.get('nombre', '').strip()
    descripcion = request.form.get('descripcion', '')
    codigo_barras = request.form.get('codigo_barras', '')
    cantidad = int(request.form.get('cantidad', 1))
    precio = request.form.get('precio_unitario')
    precio_unitario = float(precio) if precio else None
    tipo_producto_id = (request.form.get('tipo_producto_id') or '').strip() or None
    nuevo_tipo_nombre = (request.form.get('nuevo_tipo_nombre') or '').strip()
    lat = request.form.get('latitud')
    lng = request.form.get('longitud')
    latitud = float(lat) if lat else None
    longitud = float(lng) if lng else None

    if not nombre:
        return jsonify({'error': 'El nombre del producto es obligatorio'}), 400

    if nuevo_tipo_nombre and not tipo_producto_id:
        conn = get_db()
        cursor = conn.cursor()
        tipo_id = f"tipo-{uuid.uuid4().hex[:8]}"
        cursor.execute(
            "INSERT INTO tipos_producto (id, nombre, descripcion, icono, color) VALUES (?, ?, ?, ?, ?)",
            (tipo_id, nuevo_tipo_nombre, 'Creado desde app', '📦', '#3b82f6')
        )
        conn.commit()
        tipo_producto_id = tipo_id
        conn.close()

    if not tipo_producto_id:
        return jsonify({'error': 'Debe seleccionar un tipo'}), 400

    foto_url = None
    foto_thumbnail = None
    if 'foto' in request.files:
        file = request.files['foto']
        if file and file.filename:
            foto_url, foto_thumbnail = guardar_foto(file.read(), file.mimetype)
    elif request.form.get('foto_base64'):
        foto_b64 = request.form.get('foto_base64')
        if ',' in foto_b64:
            foto_b64 = foto_b64.split(',')[1]
        img_data = base64.b64decode(foto_b64)
        foto_url, foto_thumbnail = guardar_foto(img_data)

    producto_id = f"prod-{uuid.uuid4().hex[:8]}"
    ahora = datetime.now().isoformat()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO productos (id, nombre, descripcion, codigo_barras, cantidad, precio_unitario,
        tipo_producto_id, foto_url, foto_thumbnail, texto_ocr, latitud, longitud,
        fecha_creacion, fecha_actualizacion, sincronizado)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (producto_id, nombre, descripcion, codigo_barras, cantidad, precio_unitario,
          tipo_producto_id, foto_url, foto_thumbnail, '', latitud, longitud, ahora, ahora, 1))
    conn.commit()
    cursor.execute("""
        SELECT p.*, t.nombre as tipo_nombre, t.icono as tipo_icono, t.color as tipo_color
        FROM productos p
        JOIN tipos_producto t ON p.tipo_producto_id = t.id
        WHERE p.id = ?
    """, (producto_id,))
    producto = dict(cursor.fetchone())
    conn.close()
    return jsonify(producto), 201


@api_bp.put('/api/productos/<producto_id>')
def update_producto(producto_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE id = ?", (producto_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'No encontrado'}), 404
    actual = dict(row)
    conn.close()

    data = request.form if request.form else (request.get_json(silent=True) or {})

    def campo(clave, defecto):
        valor = data.get(clave)
        return valor if valor is not None else defecto

    nombre = (campo('nombre', actual['nombre']) or '').strip()
    if not nombre:
        return jsonify({'error': 'El nombre del producto es obligatorio'}), 400

    descripcion = campo('descripcion', actual['descripcion'])
    codigo_barras = campo('codigo_barras', actual['codigo_barras'])

    try:
        cantidad = int(campo('cantidad', actual['cantidad']) or 1)
        precio_raw = data.get('precio_unitario')
        if precio_raw is None:
            precio_unitario = actual['precio_unitario']
        else:
            precio_unitario = float(precio_raw) if str(precio_raw).strip() else None
    except (ValueError, TypeError):
        return jsonify({'error': 'Cantidad o precio inválido'}), 400

    tipo_producto_id = (campo('tipo_producto_id', actual['tipo_producto_id']) or '').strip()
    if not tipo_producto_id:
        return jsonify({'error': 'Debe seleccionar un tipo'}), 400

    foto_url = actual['foto_url']
    foto_thumbnail = actual['foto_thumbnail']
    nueva_foto = None
    if 'foto' in request.files and request.files['foto'].filename:
        file = request.files['foto']
        nueva_foto = (file.read(), file.mimetype)
    elif data.get('foto_base64'):
        foto_b64 = data.get('foto_base64')
        if ',' in foto_b64:
            foto_b64 = foto_b64.split(',')[1]
        nueva_foto = (base64.b64decode(foto_b64), 'image/jpeg')

    conn = get_db()
    cursor = conn.cursor()
    if nueva_foto:
        foto_url, foto_thumbnail = guardar_foto(*nueva_foto)
        if actual['foto_url'] and actual['foto_url'].startswith('/fotos/'):
            cursor.execute("DELETE FROM fotos WHERE id = ?", (actual['foto_url'].rsplit('/', 1)[1],))

    cursor.execute("""
        UPDATE productos SET nombre = ?, descripcion = ?, codigo_barras = ?, cantidad = ?,
        precio_unitario = ?, tipo_producto_id = ?, foto_url = ?, foto_thumbnail = ?,
        fecha_actualizacion = ?
        WHERE id = ?
    """, (nombre, descripcion, codigo_barras, cantidad, precio_unitario,
          tipo_producto_id, foto_url, foto_thumbnail, datetime.now().isoformat(), producto_id))
    conn.commit()
    cursor.execute("""
        SELECT p.*, t.nombre as tipo_nombre, t.icono as tipo_icono, t.color as tipo_color
        FROM productos p
        JOIN tipos_producto t ON p.tipo_producto_id = t.id
        WHERE p.id = ?
    """, (producto_id,))
    producto = dict(cursor.fetchone())
    conn.close()
    return jsonify(producto), 200


@api_bp.delete('/api/productos/<producto_id>')
def delete_producto(producto_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT foto_url FROM productos WHERE id = ?", (producto_id,))
    row = cursor.fetchone()

    if row and row['foto_url']:
        foto_url = row['foto_url']
        if foto_url.startswith('/fotos/'):
            cursor.execute("DELETE FROM fotos WHERE id = ?", (foto_url.rsplit('/', 1)[1],))
        else:
            foto_path = foto_url.replace('/uploads/fotos/', '')
            full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], foto_path)
            if os.path.exists(full_path):
                os.remove(full_path)

    cursor.execute("DELETE FROM productos WHERE id = ?", (producto_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Eliminado'}), 200


@api_bp.post('/api/clasificar')
def clasificar_producto():
    if not gemini_client:
        return jsonify({'error': 'Clasificación no configurada'}), 503

    data = request.get_json(silent=True) or {}
    foto_b64 = data.get('foto_base64', '')
    if not foto_b64:
        return jsonify({'error': 'Foto requerida'}), 400
    if ',' in foto_b64:
        foto_b64 = foto_b64.split(',')[1]

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre FROM tipos_producto ORDER BY nombre")
    tipos = [dict(row) for row in cursor.fetchall()]
    conn.close()

    nombres_tipos = [t['nombre'] for t in tipos]
    mapa_tipos = {t['nombre'].lower(): t['id'] for t in tipos}
    prompt = (
        f"Analiza esta imagen de un producto. "
        f"Clasifícalo en UNA de estas categorías: {', '.join(nombres_tipos)}. "
        f"También sugiere un nombre corto para el producto. "
        f"Responde SOLO con JSON válido, sin markdown, con este formato exacto: "
        f'{{"categoria": "nombre_categoria", "nombre_sugerido": "nombre_producto"}}'
    )

    try:
        from google.genai import types

        img_bytes = base64.b64decode(foto_b64)
        response = gemini_client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"),
                prompt,
            ],
        )
        texto = response.text.strip()
        if texto.startswith('```'):
            texto = texto.split('\n', 1)[1].rsplit('```', 1)[0].strip()
        resultado = json.loads(texto)
        categoria = resultado.get('categoria', '')
        tipo_id = mapa_tipos.get(categoria.lower())
        return jsonify({
            'tipo_id': tipo_id,
            'tipo_nombre': categoria,
            'nombre_sugerido': resultado.get('nombre_sugerido', ''),
        })
    except Exception as e:
        print(f"Error clasificando: {e}")
        return jsonify({'error': 'No se pudo clasificar'}), 500


@api_bp.get('/api/estadisticas')
def get_estadisticas():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as count FROM productos")
    total_productos = cursor.fetchone()['count']
    cursor.execute("SELECT COALESCE(SUM(cantidad * COALESCE(precio_unitario, 0)), 0) as valor FROM productos")
    valor_total = cursor.fetchone()['valor']
    cursor.execute("""
        SELECT t.nombre, t.icono, t.color, COUNT(p.id) as cantidad
        FROM tipos_producto t
        LEFT JOIN productos p ON t.id = p.tipo_producto_id
        GROUP BY t.id
        ORDER BY cantidad DESC
    """)
    por_tipo = [dict(row) for row in cursor.fetchall()]
    hace_7_dias = (datetime.now() - timedelta(days=7)).isoformat()
    cursor.execute("SELECT COUNT(*) as count FROM productos WHERE fecha_creacion >= ?", (hace_7_dias,))
    recientes = cursor.fetchone()['count']
    conn.close()
    return jsonify({
        'total_productos': total_productos,
        'valor_total': round(valor_total, 2),
        'productos_recientes': recientes,
        'por_tipo': por_tipo
    })


def _productos_para_export():
    tipo_id = request.args.get('tipo_id')
    busqueda = request.args.get('q')
    conn = get_db()
    cursor = conn.cursor()
    query = """
        SELECT p.*, t.nombre as tipo_nombre
        FROM productos p
        JOIN tipos_producto t ON p.tipo_producto_id = t.id
        WHERE 1=1
    """
    params = []
    if tipo_id:
        query += " AND p.tipo_producto_id = ?"
        params.append(tipo_id)
    if busqueda:
        query += " AND (p.nombre LIKE ? OR p.codigo_barras LIKE ?)"
        params.extend([f'%{busqueda}%', f'%{busqueda}%'])
    query += " ORDER BY p.fecha_creacion DESC"
    cursor.execute(query, params)
    productos = []
    for row in cursor.fetchall():
        p = dict(row)
        cantidad = p.get('cantidad') or 0
        precio = p.get('precio_unitario')
        p['valor_total'] = round(cantidad * precio, 2) if precio is not None else None
        productos.append(p)
    conn.close()
    return productos


@api_bp.get('/api/export')
def export_inventario():
    formato = request.args.get('formato', 'xlsx').lower()
    productos = _productos_para_export()
    fecha = datetime.now().strftime('%Y-%m-%d')

    if formato == 'csv':
        import csv
        salida = io.StringIO()
        writer = csv.writer(salida)
        writer.writerow([titulo for _, titulo in EXPORT_COLUMNAS])
        for p in productos:
            writer.writerow([p.get(clave, '') for clave, _ in EXPORT_COLUMNAS])
        contenido = chr(0xFEFF) + salida.getvalue()
        return Response(
            contenido,
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename=inventario_{fecha}.csv'}
        )

    if formato == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventario'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        for col, (_, titulo) in enumerate(EXPORT_COLUMNAS, start=1):
            celda = ws.cell(row=1, column=col, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')
        for fila, p in enumerate(productos, start=2):
            for col, (clave, _) in enumerate(EXPORT_COLUMNAS, start=1):
                ws.cell(row=fila, column=col, value=p.get(clave))
        fila_total = len(productos) + 2
        ws.cell(row=fila_total, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=fila_total, column=3, value=sum((p.get('cantidad') or 0) for p in productos)).font = Font(bold=True)
        ws.cell(row=fila_total, column=5, value=round(sum((p.get('valor_total') or 0) for p in productos), 2)).font = Font(bold=True)
        anchos = [30, 16, 10, 14, 12, 18, 30, 12, 12, 20]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(EXPORT_COLUMNAS))}{len(productos) + 1}'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=inventario_{fecha}.xlsx'}
        )

    return jsonify({'error': "Formato no soportado (usa 'xlsx' o 'csv')"}), 400


@api_bp.post('/api/sync')
def sync_offline():
    data = request.get_json(silent=True) or {}
    productos_pendientes = data.get('productos', [])
    sincronizados = []

    for item in productos_pendientes:
        try:
            nombre = (item.get('nombre') or '').strip()
            if not nombre:
                continue

            producto_id = f"prod-{uuid.uuid4().hex[:8]}"
            foto_url = None
            foto_thumbnail = None
            if item.get('foto_base64'):
                foto_b64 = item['foto_base64']
                if ',' in foto_b64:
                    foto_b64 = foto_b64.split(',')[1]
                img_data = base64.b64decode(foto_b64)
                foto_url, foto_thumbnail = guardar_foto(img_data)

            tipo_id = item.get('tipo_producto_id')
            if tipo_id:
                tipo_id = str(tipo_id).strip() or None
            else:
                tipo_id = None

            nuevo = (item.get('nuevo_tipo_nombre') or '').strip()
            if nuevo and not tipo_id:
                tipo_id = f"tipo-{uuid.uuid4().hex[:8]}"
                conn = get_db()
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO tipos_producto (id, nombre, descripcion, icono, color) VALUES (?, ?, ?, ?, ?)",
                    (tipo_id, nuevo, 'Creado offline', '📦', '#3b82f6')
                )
                conn.commit()
                conn.close()

            if not tipo_id:
                continue

            ahora = datetime.now().isoformat()
            sync_lat = float(item['latitud']) if item.get('latitud') else None
            sync_lng = float(item['longitud']) if item.get('longitud') else None
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO productos (id, nombre, descripcion, codigo_barras, cantidad, precio_unitario,
                tipo_producto_id, foto_url, foto_thumbnail, texto_ocr, latitud, longitud,
                fecha_creacion, fecha_actualizacion, sincronizado)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (producto_id, nombre, item.get('descripcion', ''), item.get('codigo_barras', ''),
                  item.get('cantidad', 1), item.get('precio_unitario'), tipo_id, foto_url, foto_thumbnail,
                  item.get('texto_ocr', ''), sync_lat, sync_lng, item.get('fecha_creacion', ahora), ahora, 1))
            conn.commit()
            conn.close()
            sincronizados.append({'temp_id': item.get('temp_id'), 'server_id': producto_id})
        except Exception as e:
            print(f"Error sync: {e}")

    return jsonify({'sincronizados': len(sincronizados), 'detalles': sincronizados})
